# coding UTF-8
# author:huangZengli
from osgeo import gdal
from pylab import *  # 支持中文
import time
import os
import pandas as pd
from time import strftime, gmtime

mpl.rcParams['font.sans-serif'] = ['SimHei']
from openpyxl import Workbook
 
# 创建一个Workbook对象
work = Workbook()
 
def out(data, name):
    ws = work.active
    ws['A1'] = '经度'
    ws['B1'] = '纬度'
    ws['C1'] = '高程'
    ws['D1'] = '所在栅格行'
    ws['E1'] = '所在栅格列'
    for i in range(len(data)):
        rows = []
        row_length = len(data[i])
        #if row_length != 0:
        for j in range(row_length):
            rows.append(data[i][j])
            ws.append(rows[j])
        #print(rows)
    work.save(name)
 
if __name__ == "__main__":
    import argparse
     
    parser = argparse.ArgumentParser(description='Trans latlot')
    parser.add_argument('--task_path',
                    default='C:/lyd_software/0.pythonCode/lyd_utils/task',
                    help='root dir')

    args = parser.parse_args()
    start = time.time()
    print('开始转换')
    # 单文件
    # filePath = 'C://Users//V//Documents//DJI//DJITerra//wcl1991@qq.com//save//save48.tif'  # tif文件路径
    
    # 文件列表
    filedir = args.task_path
    files = os.listdir(filedir+"/patch_save/") #得到文件夹下的所有文件名称
    files.sort(key=lambda x:int(x[:-4])) # 按顺序执行文件

    for file in files: #遍历文件夹
     if not os.path.isdir(file): #判断是否是文件夹，不是文件夹才打开
           
        dataset = gdal.Open(filedir+"/patch_save/"+file)  # 打开tif
        print("正在处理文件："+str(file))
        # 获取行数列数和地理信息
        # geo_information(0):左上像素左上角的x坐标。
        # geo_information(1):w - e像素分辨率 / 像素宽度。
        # geo_information(2):行旋转（通常为零）。
        # geo_information(3):左上像素左上角的y坐标。
        # geo_information(4):列旋转（通常为零）。
        # geo_information(5):n - s像素分辨率 / 像素高度（北半球上图像为负值）
        geo_information = dataset.GetGeoTransform()
        col = dataset.RasterXSize  # 29602
        row = dataset.RasterYSize  # 1000
        #print(col, row)

        band = dataset.RasterCount
        dem = dataset.GetRasterBand(1).ReadAsArray()

        # 记录每个文件的行数据，集成一个文件
        cols = []
        for y in range(row):  # 行
            rows = []
            for x in range(col):  # 列
                    # 输出经纬度
                    lon = geo_information[0] + x * geo_information[1] + y * geo_information[2]
                    lat = geo_information[3] + x * geo_information[4] + y * geo_information[5]
                    #child = [lon, lat, dem[y][x], y+img_height*yi, x]
                    # 经纬度，高程，行，列
                    child = [lon, lat, dem[y][x], y, x]
                    #print(child)
                    rows.append(child)
            cols.append(rows)
            del rows
        # 删除文件
        del dataset
        del geo_information
        print("---------"+str(file)+"---------Done")

        # 输出xls文件
    #out(cols, filedir+"//xls//"+"result.xlsx")
        #输出csv文件
        data = pd.DataFrame(cols)
        data.to_csv(filedir+"//dictionary//"+file[0: -4]+".csv")
        print(str(file)+"字典已经生成")
    end = time.time()
    print ("耗时："+strftime("%H:%M:%S", gmtime(end - start)))